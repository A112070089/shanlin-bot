import os
import json
import hmac
import hashlib
import base64
from datetime import datetime

from fastapi import FastAPI, Request, HTTPException, Header, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from pydantic import BaseModel
from typing import Optional
import io

import httpx
import firebase_admin
from firebase_admin import credentials, db
from openpyxl import load_workbook

app = FastAPI(title="山林診所 LINE Bot API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

CHANNEL_SECRET       = os.environ.get("CHANNEL_SECRET", "")
CHANNEL_ACCESS_TOKEN = os.environ.get("CHANNEL_ACCESS_TOKEN", "")
FIREBASE_URL         = os.environ.get("FIREBASE_URL", "")
LIFF_VERIFY_URL      = os.environ.get("LIFF_URL", "https://liff.line.me/2010169963-KEjAbfsW")
LIFF_APPT_URL        = os.environ.get("LIFF_APPT_URL", "https://liff.line.me/2010169963-n8zZXE1V")
ADMIN_PASSWORD       = os.environ.get("ADMIN_PASSWORD", "shanlin2025")
LINE_API             = "https://api.line.me/v2/bot/message"

firebase_cred_json = os.environ.get("FIREBASE_CREDENTIALS", "")
try:
    if firebase_cred_json and firebase_cred_json != "{}" and not firebase_admin._apps:
        cred_dict = json.loads(firebase_cred_json)
        if cred_dict.get("type") == "service_account":
            cred = credentials.Certificate(cred_dict)
            firebase_admin.initialize_app(cred, {"databaseURL": FIREBASE_URL})
except Exception as e:
    print(f"Firebase init skipped: {e}")


def verify_signature(body: bytes, signature: str) -> bool:
    hash_ = hmac.new(CHANNEL_SECRET.encode(), body, hashlib.sha256).digest()
    expected = base64.b64encode(hash_).decode()
    return hmac.compare_digest(expected, signature)


async def push_message(user_id: str, messages: list):
    async with httpx.AsyncClient() as client:
        res = await client.post(
            f"{LINE_API}/push",
            headers={"Authorization": f"Bearer {CHANNEL_ACCESS_TOKEN}"},
            json={"to": user_id, "messages": messages}
        )
        print(f"Push status: {res.status_code}, body: {res.text}")


async def reply_message(reply_token: str, messages: list):
    async with httpx.AsyncClient() as client:
        res = await client.post(
            f"{LINE_API}/reply",
            headers={"Authorization": f"Bearer {CHANNEL_ACCESS_TOKEN}"},
            json={"replyToken": reply_token, "messages": messages}
        )
        print(f"Reply status: {res.status_code}, body: {res.text}")


def make_liff_button(label: str, url: str) -> dict:
    return {
        "type": "template",
        "altText": label,
        "template": {
            "type": "buttons",
            "text": label,
            "actions": [{
                "type": "uri",
                "label": label[:20],
                "uri": url
            }]
        }
    }


def check_admin(x_admin_password: Optional[str]):
    if x_admin_password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="密碼錯誤")


# ══════════════════════════════════════
#  LINE Webhook
# ══════════════════════════════════════
@app.post("/webhook")
async def webhook(request: Request):
    body = await request.body()
    signature = request.headers.get("X-Line-Signature", "")

    if CHANNEL_SECRET and not verify_signature(body, signature):
        raise HTTPException(status_code=400, detail="Invalid signature")

    data = json.loads(body)

    for event in data.get("events", []):
        event_type  = event.get("type")
        user_id     = event.get("source", {}).get("userId", "")
        reply_token = event.get("replyToken", "")

        if event_type == "follow":
            await handle_follow(user_id, reply_token)
        elif event_type == "message" and event["message"]["type"] == "text":
            text = event["message"]["text"].strip()
            await handle_text(user_id, reply_token, text)

    return JSONResponse({"status": "ok"})


async def handle_follow(user_id: str, reply_token: str):
    await reply_message(reply_token, [
        {
            "type": "text",
            "text": (
                "感謝您成為山林診所好友 💗\n\n"
                "⭐ 門診時間\n"
                "週一至週五 早上 8:30-12:30、下午 13:30-17:30\n"
                "週六 早上 8:30-12:30\n"
                "週日公休\n\n"
                "⭐ 診所位置：台北市文山區羅斯福路六段 407 號 2 樓\n"
                "⭐ 連絡電話：02-2933-2010\n\n"
                "請使用下方選單選擇服務 👇"
            )
        }
    ])


async def handle_text(user_id: str, reply_token: str, text: str):
    keywords_booking = ["預約", "體檢", "健檢", "掛號"]
    keywords_verify  = ["綁定", "驗證", "身分證"]
    keywords_query   = ["我的預約", "查詢預約", "預約資料", "預約紀錄", "查預約"]
    keywords_report  = ["報告", "檢查結果"]
    keywords_hours   = ["時間", "門診", "幾點"]
    keywords_address = ["地址", "位置", "在哪", "怎麼去"]

    if any(k in text for k in keywords_query):
        appt_data = None
        try:
            all_appts = db.reference("appointments").get() or {}
            for id_num, data in all_appts.items():
                if isinstance(data, dict) and data.get("lineUserId") == user_id:
                    appt_data = data
                    break
        except Exception as e:
            print(f"query error: {e}")

        if appt_data:
            plan_names = {
                "A": "A 方案（腦肺方案）",
                "B": "B 方案（腹部方案）",
                "C": "C 方案（骨密肌力方案）"
            }
            await reply_message(reply_token, [{
                "type": "text",
                "text": (
                    f"📋 您的預約資料\n\n"
                    f"👤 姓名：{appt_data.get('name', '—')}\n"
                    f"📅 日期：{appt_data.get('date', '—')}\n"
                    f"⏰ 時段：{appt_data.get('time', '—')}\n"
                    f"🏥 方案：{plan_names.get(appt_data.get('plan', ''), appt_data.get('plan', '—'))}\n\n"
                    f"如需修改請撥打 02-2933-2010"
                )
            }])
        else:
            await reply_message(reply_token, [{
                "type": "text",
                "text": "目前查無您的預約紀錄。\n如需預約請說「預約」，或撥打 02-2933-2010 😊"
            }])

    elif any(k in text for k in keywords_booking):
        await reply_message(reply_token, [
            {
                "type": "text",
                "text": "您好！請點下方按鈕進行健檢預約 📋"
            },
            make_liff_button("📅 立即預約健檢", LIFF_APPT_URL)
        ])

    elif any(k in text for k in keywords_verify):
        await reply_message(reply_token, [
            {
                "type": "text",
                "text": (
                    "如果您已在診所完成預約，\n"
                    "請點下方按鈕綁定 LINE 帳號，\n"
                    "之後即可收到禁食提醒與報告通知 🔔"
                )
            },
            make_liff_button("🔐 綁定 LINE 帳號", LIFF_VERIFY_URL)
        ])

    elif any(k in text for k in keywords_report):
        await reply_message(reply_token, [
            {
                "type": "text",
                "text": "您的健檢報告可在綁定後查閱，AI 會用白話文解讀紅字數值 📋"
            },
            make_liff_button("📄 查看我的報告", LIFF_VERIFY_URL)
        ])

    elif any(k in text for k in keywords_hours):
        await reply_message(reply_token, [{
            "type": "text",
            "text": (
                "⭐ 門診時間\n"
                "週一至週五 早上 8:30-12:30、下午 13:30-17:30\n"
                "週六 早上 8:30-12:30\n"
                "週日公休\n"
                "國定假日門診半天"
            )
        }])

    elif any(k in text for k in keywords_address):
        await reply_message(reply_token, [{
            "type": "text",
            "text": "📍 台北市文山區羅斯福路六段 407 號 2 樓\n（由車前路門口上樓）\n\n📞 02-2933-2010"
        }])

    else:
        await reply_message(reply_token, [{
            "type": "text",
            "text": "感謝您的訊息！如需協助請使用下方選單，或撥打 02-2933-2010 🙏"
        }])


# ══════════════════════════════════════
#  LIFF 身分驗證綁定 API
# ══════════════════════════════════════
class VerifyRequest(BaseModel):
    id_number:    str
    phone:        str
    line_user_id: str


@app.post("/api/liff/verify")
async def liff_verify(req: VerifyRequest):
    id_num = req.id_number.upper().strip()
    phone  = req.phone.strip()

    try:
        ref  = db.reference(f"appointments/{id_num}")
        data = ref.get()
    except Exception:
        raise HTTPException(status_code=404, detail="查無資料")

    if not data:
        raise HTTPException(status_code=404, detail="查無此身分證資料")

    stored_phone = str(data.get("phone", ""))
    if not stored_phone.startswith("0"):
        stored_phone = "0" + stored_phone
    if stored_phone != phone:
        raise HTTPException(status_code=401, detail="手機號碼不符")

    ref.update({"lineUserId": req.line_user_id, "boundAt": datetime.now().isoformat()})

    await push_message(req.line_user_id, [
        {
            "type": "text",
            "text": (
                f"✅ LINE 帳號綁定成功！\n\n"
                f"您好，{data.get('name', '')}！\n"
                f"之後健檢前我們會主動提醒您禁食注意事項，\n"
                f"健檢完成後也可以在 LINE 查看報告 📋"
            )
        }
    ])

    return {
        "name": data.get("name"),
        "date": data.get("date"),
        "time": data.get("time"),
        "plan": data.get("plan"),
    }


# ══════════════════════════════════════
#  語音預約 API
# ══════════════════════════════════════
class AppointmentRequest(BaseModel):
    name:         str
    phone:        str
    plan:         str
    date:         str
    time_slot:    str
    line_user_id: str = ""
    birth:        str = ""
    id_number:    str = ""


@app.post("/api/appointment/book")
async def book_appointment(req: AppointmentRequest):
    plan_names = {
        "A": "A 方案（腦肺方案）",
        "B": "B 方案（腹部方案）",
        "C": "C 方案（骨密肌力方案）"
    }

    phone = req.phone.strip()
    if not phone.startswith("0"):
        phone = "0" + phone

    try:
        all_appts = db.reference("appointments").get() or {}
        target_key = None
        for id_num, data in all_appts.items():
            if isinstance(data, dict):
                stored = str(data.get("phone", ""))
                if not stored.startswith("0"):
                    stored = "0" + stored
                if stored == phone:
                    target_key = id_num
                    break

        record = {
            "name":     req.name,
            "phone":    phone,
            "plan":     req.plan,
            "date":     req.date,
            "time":     req.time_slot,
            "bookedAt": datetime.now().isoformat(),
            "source":   "LINE",
        }
        if req.birth:
            record["birth"] = req.birth
        if req.id_number:
            record["idNumber"] = req.id_number.upper()
        if req.line_user_id:
            record["lineUserId"] = req.line_user_id

        if target_key:
            db.reference(f"appointments/{target_key}").update(record)
        else:
            new_key = req.id_number.upper().strip() if req.id_number else f"LIFF_{datetime.now().strftime('%Y%m%d%H%M%S')}"
            db.reference(f"appointments/{new_key}").set(record)

    except Exception as e:
        print(f"book error: {e}")

    if req.line_user_id:
        await push_message(req.line_user_id, [{
            "type": "text",
            "text": (
                f"✅ 預約成功！\n\n"
                f"📋 方案：{plan_names.get(req.plan, req.plan)}\n"
                f"📅 日期：{req.date}\n"
                f"⏰ 時段：{req.time_slot}\n\n"
                f"健檢前我們會提醒您禁食注意事項 😊"
            )
        }])

    return {"status": "success", "message": "預約成功"}


# ══════════════════════════════════════
#  診前提醒
# ══════════════════════════════════════
@app.post("/api/reminder/send")
async def send_reminders():
    import datetime as dt
    tomorrow = (dt.date.today() + dt.timedelta(days=1)).isoformat()

    try:
        all_appts = db.reference("appointments").get() or {}
    except Exception:
        return {"status": "error", "message": "Firebase 連線失敗"}

    sent = 0
    for id_num, data in all_appts.items():
        if not isinstance(data, dict):
            continue
        if data.get("date") == tomorrow and data.get("lineUserId"):
            await push_message(data["lineUserId"], [{
                "type": "text",
                "text": (
                    f"⏰ 健檢提醒\n\n"
                    f"{data.get('name', '您好')}，明天 {tomorrow} 您有健檢預約！\n\n"
                    f"📌 注意事項：\n"
                    f"• 今晚 10 點後請禁食禁水\n"
                    f"• 請攜帶健保卡與身分證\n"
                    f"• 穿著輕便衣物\n\n"
                    f"健檢時間：{data.get('time', '')}，請準時到達 🏥"
                )
            }])
            sent += 1

    return {"status": "success", "sent": sent}


# ══════════════════════════════════════
#  後台管理 API
# ══════════════════════════════════════
@app.post("/api/admin/login")
async def admin_login(payload: dict):
    password = payload.get("password", "")
    if password != ADMIN_PASSWORD:
        raise HTTPException(status_code=401, detail="密碼錯誤")
    return {"status": "ok"}


@app.get("/api/admin/appointments")
async def admin_list_appointments(x_admin_password: Optional[str] = Header(None)):
    check_admin(x_admin_password)
    try:
        all_appts = db.reference("appointments").get() or {}
    except Exception:
        return {"appointments": []}

    result = []
    for key, data in all_appts.items():
        if isinstance(data, dict):
            item = dict(data)
            item["_key"] = key
            item["_bound"] = bool(data.get("lineUserId"))
            result.append(item)

    result.sort(key=lambda x: x.get("date", ""), reverse=True)
    return {"appointments": result}


class AdminCreateRequest(BaseModel):
    name:      str
    phone:     str
    plan:      str
    date:      str
    time:      str
    source:    str = "電話"
    birth:     str = ""
    id_number: str = ""
    note:      str = ""


@app.post("/api/admin/appointments")
async def admin_create_appointment(req: AdminCreateRequest, x_admin_password: Optional[str] = Header(None)):
    check_admin(x_admin_password)

    phone = req.phone.strip()
    if phone and not phone.startswith("0"):
        phone = "0" + phone

    record = {
        "name":     req.name,
        "phone":    phone,
        "plan":     req.plan,
        "date":     req.date,
        "time":     req.time,
        "source":   req.source,
        "bookedAt": datetime.now().isoformat(),
    }
    if req.birth:
        record["birth"] = req.birth
    if req.id_number:
        record["idNumber"] = req.id_number.upper()
    if req.note:
        record["note"] = req.note

    key = req.id_number.upper().strip() if req.id_number else f"ADMIN_{datetime.now().strftime('%Y%m%d%H%M%S')}"

    try:
        db.reference(f"appointments/{key}").set(record)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    return {"status": "success", "key": key}


class AdminUpdateRequest(BaseModel):
    name:      Optional[str] = None
    phone:     Optional[str] = None
    plan:      Optional[str] = None
    date:      Optional[str] = None
    time:      Optional[str] = None
    source:    Optional[str] = None
    birth:     Optional[str] = None
    id_number: Optional[str] = None
    note:      Optional[str] = None


@app.put("/api/admin/appointments/{key}")
async def admin_update_appointment(key: str, req: AdminUpdateRequest, x_admin_password: Optional[str] = Header(None)):
    check_admin(x_admin_password)

    update_data = {}
    if req.name is not None: update_data["name"] = req.name
    if req.phone is not None:
        p = req.phone.strip()
        if p and not p.startswith("0"):
            p = "0" + p
        update_data["phone"] = p
    if req.plan is not None: update_data["plan"] = req.plan
    if req.date is not None: update_data["date"] = req.date
    if req.time is not None: update_data["time"] = req.time
    if req.source is not None: update_data["source"] = req.source
    if req.birth is not None: update_data["birth"] = req.birth
    if req.id_number is not None: update_data["idNumber"] = req.id_number.upper()
    if req.note is not None: update_data["note"] = req.note

    try:
        db.reference(f"appointments/{key}").update(update_data)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    return {"status": "success"}


@app.delete("/api/admin/appointments/{key}")
async def admin_delete_appointment(key: str, x_admin_password: Optional[str] = Header(None)):
    check_admin(x_admin_password)
    try:
        db.reference(f"appointments/{key}").delete()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    return {"status": "success"}


# ══════════════════════════════════════
#  Excel 批次匯入 API
# ══════════════════════════════════════
@app.post("/api/admin/import")
async def admin_import_excel(file: UploadFile = File(...), x_admin_password: Optional[str] = Header(None)):
    check_admin(x_admin_password)

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="請上傳 Excel 檔案（.xlsx 或 .xls）")

    contents = await file.read()
    try:
        wb = load_workbook(io.BytesIO(contents), data_only=True)
        sheet = wb.worksheets[0]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"無法讀取 Excel 檔案：{e}")

    rows = list(sheet.iter_rows(min_row=2, values_only=True))

    imported = 0
    skipped = 0
    errors = []

    for idx, row in enumerate(rows, start=2):
        if not row or not row[0]:
            continue

        try:
            name      = str(row[0]).strip() if row[0] else ""
            phone     = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            id_number = str(row[2]).strip().upper() if len(row) > 2 and row[2] else ""
            birth     = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            plan      = str(row[4]).strip().upper() if len(row) > 4 and row[4] else ""
            date_val  = row[5] if len(row) > 5 else None
            time_val  = row[6] if len(row) > 6 else None
            source    = str(row[7]).strip() if len(row) > 7 and row[7] else "Excel匯入"
            note      = str(row[8]).strip() if len(row) > 8 and row[8] else ""

            if not name or not phone or plan not in ("A", "B", "C"):
                skipped += 1
                errors.append(f"第 {idx} 列：缺少必填欄位或方案格式錯誤")
                continue

            # 處理日期格式（可能是 datetime 物件或字串）
            if hasattr(date_val, "strftime"):
                date_str = date_val.strftime("%Y-%m-%d")
            else:
                date_str = str(date_val).strip() if date_val else ""

            if hasattr(time_val, "strftime"):
                time_str = time_val.strftime("%H:%M")
            else:
                time_str = str(time_val).strip() if time_val else ""

            if not phone.startswith("0"):
                phone = "0" + phone

            record = {
                "name":     name,
                "phone":    phone,
                "plan":     plan,
                "date":     date_str,
                "time":     time_str,
                "source":   source,
                "bookedAt": datetime.now().isoformat(),
            }
            if birth:
                record["birth"] = birth
            if id_number:
                record["idNumber"] = id_number
            if note:
                record["note"] = note

            key = id_number if id_number else f"IMPORT_{datetime.now().strftime('%Y%m%d%H%M%S')}_{idx}"
            db.reference(f"appointments/{key}").set(record)
            imported += 1

        except Exception as e:
            skipped += 1
            errors.append(f"第 {idx} 列：{str(e)}")

    return {
        "status": "success",
        "imported": imported,
        "skipped": skipped,
        "errors": errors[:20]
    }


@app.get("/")
async def root():
    return {"status": "ok", "service": "山林診所 LINE Bot API"}
